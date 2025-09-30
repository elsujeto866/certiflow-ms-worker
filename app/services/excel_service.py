"""
Servicio para generar archivos Excel basados en plantillas.
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Any, Optional
import os
from pathlib import Path
from datetime import datetime
from app.core.config import settings
from app.models.exceptions import ExcelGenerationError, TemplateNotFoundError
from app.core.logging import get_logger

logger = get_logger(__name__)


class ExcelService:
    """Servicio para manejar la generación de archivos Excel."""
    
    def __init__(self):
        self.templates_dir = Path(settings.templates_dir)
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_excel_report(self, data: Dict[str, Any], template_name: Optional[str] = None) -> str:
        """
        Genera un reporte Excel basado en los datos extraídos.
        
        Args:
            data: Datos extraídos a incluir en el Excel
            template_name: Nombre de la plantilla a usar (opcional)
            
        Returns:
            str: Ruta del archivo Excel generado
            
        Raises:
            ExcelGenerationError: Si hay error generando el Excel
            TemplateNotFoundError: Si no se encuentra la plantilla
        """
        try:
            logger.info(f"Generando reporte Excel con plantilla: {template_name}")
            
            if template_name:
                workbook = self._load_template(template_name)
            else:
                workbook = self._create_default_workbook()
            
            # Llenar datos en el workbook
            self._populate_workbook(workbook, data)
            
            # Generar nombre de archivo único
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"certiflow_report_{timestamp}.xlsx"
            output_path = self.output_dir / filename
            
            # Guardar archivo
            workbook.save(output_path)
            logger.info(f"Reporte Excel generado: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            raise ExcelGenerationError(f"Error generando archivo Excel: {e}")
    
    def _load_template(self, template_name: str) -> Workbook:
        """
        Carga una plantilla Excel existente.
        
        Args:
            template_name: Nombre de la plantilla
            
        Returns:
            Workbook: Objeto workbook de openpyxl
            
        Raises:
            TemplateNotFoundError: Si no se encuentra la plantilla
        """
        template_path = self.templates_dir / f"{template_name}.xlsx"
        
        if not template_path.exists():
            raise TemplateNotFoundError(f"Plantilla no encontrada: {template_name}")
        
        try:
            return load_workbook(template_path)
        except Exception as e:
            raise ExcelGenerationError(f"Error cargando plantilla {template_name}: {e}")
    
    def _create_default_workbook(self) -> Workbook:
        """
        Crea un workbook con formato por defecto.
        
        Returns:
            Workbook: Nuevo workbook con formato básico
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Datos Extraídos"
        
        # Agregar encabezados con estilo
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ["Campo", "Valor", "Tipo", "Confianza"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        return workbook
    
    def _populate_workbook(self, workbook: Workbook, data: Dict[str, Any]) -> None:
        """
        Llena el workbook con los datos extraídos.
        
        Args:
            workbook: Workbook a llenar
            data: Datos a insertar
        """
        worksheet = workbook.active
        
        # Si existe una hoja específica para datos, usarla
        if "Datos" in workbook.sheetnames:
            worksheet = workbook["Datos"]
        elif "Sheet1" in workbook.sheetnames:
            worksheet = workbook["Sheet1"]
        
        # Agregar metadatos en la parte superior
        self._add_metadata(worksheet, data.get("_metadata", {}))
        
        # Encontrar la primera fila disponible
        start_row = self._find_start_row(worksheet)
        
        # Insertar datos de manera plana
        current_row = start_row
        for key, value in data.items():
            if key.startswith("_"):  # Saltar metadatos
                continue
            
            # Manejar diferentes tipos de datos
            if isinstance(value, dict):
                current_row = self._add_dict_data(worksheet, key, value, current_row)
            elif isinstance(value, list):
                current_row = self._add_list_data(worksheet, key, value, current_row)
            else:
                worksheet.cell(row=current_row, column=1, value=key)
                worksheet.cell(row=current_row, column=2, value=str(value))
                worksheet.cell(row=current_row, column=3, value=type(value).__name__)
                current_row += 1
    
    def _add_metadata(self, worksheet, metadata: Dict[str, Any]) -> None:
        """Agrega metadatos al worksheet."""
        if not metadata:
            return
        
        worksheet.cell(row=1, column=5, value="Metadatos del Procesamiento")
        worksheet.cell(row=1, column=5).font = Font(bold=True)
        
        row = 2
        for key, value in metadata.items():
            worksheet.cell(row=row, column=5, value=key)
            worksheet.cell(row=row, column=6, value=str(value))
            row += 1
    
    def _find_start_row(self, worksheet) -> int:
        """Encuentra la primera fila disponible para datos."""
        # Si hay encabezados, empezar después
        if worksheet.cell(row=1, column=1).value in ["Campo", "Dato", "Key"]:
            return 2
        return 1
    
    def _add_dict_data(self, worksheet, parent_key: str, data: Dict[str, Any], start_row: int) -> int:
        """Agrega datos de diccionario al worksheet."""
        current_row = start_row
        for key, value in data.items():
            full_key = f"{parent_key}.{key}"
            worksheet.cell(row=current_row, column=1, value=full_key)
            worksheet.cell(row=current_row, column=2, value=str(value))
            worksheet.cell(row=current_row, column=3, value=type(value).__name__)
            current_row += 1
        return current_row
    
    def _add_list_data(self, worksheet, parent_key: str, data: list, start_row: int) -> int:
        """Agrega datos de lista al worksheet."""
        current_row = start_row
        for i, item in enumerate(data):
            full_key = f"{parent_key}[{i}]"
            worksheet.cell(row=current_row, column=1, value=full_key)
            worksheet.cell(row=current_row, column=2, value=str(item))
            worksheet.cell(row=current_row, column=3, value=type(item).__name__)
            current_row += 1
        return current_row
    
    def list_templates(self) -> list:
        """
        Lista las plantillas disponibles.
        
        Returns:
            list: Lista de nombres de plantillas disponibles
        """
        if not self.templates_dir.exists():
            return []
        
        templates = []
        for file in self.templates_dir.glob("*.xlsx"):
            templates.append(file.stem)
        
        return templates